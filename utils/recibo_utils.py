from pathlib import Path
import json
try:
    from openpyxl import load_workbook, Workbook
except ModuleNotFoundError:
    load_workbook = None
    Workbook = None

# Intentamos tomar la ruta del historial desde config.py; si no existe, usamos una por defecto
try:
    from config import HISTORIAL_XLSX  # ej: Path("historial/recibos.xlsx")
except Exception:
    HISTORIAL_XLSX = Path("historial/recibos.xlsx")

def _asegurar_historial():
    """
    Crea el archivo de historial si no existe, con columnas estándar y
    agrega la columna opcional 'DatosJSON' si falta (para guardar el payload completo).
    """
    if Workbook is None:
        # No explotamos al importar el módulo; explicamos al usarlo
        raise RuntimeError("Falta la dependencia 'openpyxl'. Instalá con: pip install -r RecibosDigitales/requirements.txt")
    p = Path(HISTORIAL_XLSX)
    p.parent.mkdir(parents=True, exist_ok=True)
    if not p.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Recibos"
        ws.append(["Número", "Cliente", "Fecha", "Subtotal", "Total", "Estado", "DatosJSON"])
        wb.save(p)
        return

    # Si existe, asegurar que tenga la columna 'DatosJSON'
    wb = load_workbook(p)
    ws = wb.active
    headers = [c.value for c in ws[1]] if ws.max_row >= 1 else []
    if headers and (len(headers) < 7 or headers[6] != "DatosJSON"):
        # Agregar encabezado en columna G (7)
        ws.cell(row=1, column=7, value="DatosJSON")
        wb.save(p)
def posible_duplicado(cliente: str, fecha: str, total: float) -> bool:
    """
    True si ya hay una fila con el mismo (cliente, fecha, total).
    """
    _asegurar_historial()
    wb = load_workbook(HISTORIAL_XLSX)
    ws = wb.active
    cli_norm = (cliente or "").strip().lower()
    for row in ws.iter_rows(min_row=2, values_only=True):
        _num, _cli, _fec, _sub, _tot, _est = row
        try:
            tot_val = float(str(_tot).replace(",", "."))
        except Exception:
            tot_val = None
        if (str(_fec) == fecha
            and (str(_cli or "").strip().lower() == cli_norm)
            and tot_val is not None
            and abs(tot_val - float(total)) < 0.01):
            return True
    return False

def marcar_anulado(numero: str):
    """
    Marca como 'Anulado' el recibo con ese número en el historial.
    Si no existe, agrega una fila nueva con Estado = 'Anulado'.
    """
    _asegurar_historial()
    wb = load_workbook(HISTORIAL_XLSX)
    ws = wb.active

    encontrado = False
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(numero):
            # Columna 6 = 'Estado' (índice 5)
            row[5].value = "Anulado"
            encontrado = True
            break

    if not encontrado:
        # Si no estaba en el historial, lo agregamos como anulado
        ws.append([numero, "", "", "", "", "Anulado"])

    wb.save(HISTORIAL_XLSX)

def upsert_historial_con_json(
    numero: str,
    cliente: str,
    fecha: str,
    subtotal: str | float | None,
    total: str | float | None,
    estado: str = "",
    datos: dict | None = None,
):
    """
    Inserta o actualiza la fila del historial por 'Número'.
    Guarda además 'DatosJSON' (payload completo) en la columna 7 si se provee.
    """
    _asegurar_historial()
    p = Path(HISTORIAL_XLSX)
    wb = load_workbook(p)
    ws = wb.active

    json_str = None
    if datos is not None:
        try:
            json_str = json.dumps(datos, ensure_ascii=False)
        except Exception:
            json_str = None

    encontrado = False
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(numero):
            row[1].value = cliente
            row[2].value = fecha
            row[3].value = subtotal
            row[4].value = total
            row[5].value = estado
            if json_str is not None:
                # Columna 7
                if ws.max_column < 7:
                    ws.cell(row=1, column=7, value="DatosJSON")
                ws.cell(row=row[0].row, column=7, value=json_str)
            encontrado = True
            break

    if not encontrado:
        # Asegurar encabezado JSON
        if ws.max_column < 7:
            ws.cell(row=1, column=7, value="DatosJSON")
        ws.append([numero, cliente, fecha, subtotal, total, estado, json_str])

    wb.save(p)
