import tkinter as tk
from tkinter import ttk, messagebox
from pathlib import Path
import json
import re

try:
    import openpyxl
except ModuleNotFoundError:
    openpyxl = None

from config import SALIDA_DIR, ASSETS_DIR, BASE_QR_URL, QR_SECRET_KEY
from utils.qr_utils import build_qr_data
from utils.pdf_generator import generar_pdf
from utils.recibo_utils import upsert_historial_con_json
from utils import recibos_db

# Ruta del historial (opcional desde config)
try:
    from config import HISTORIAL_XLSX as HISTORIAL_PATH
except Exception:
    HISTORIAL_PATH = Path("historial/recibos.xlsx")


def _notify(notify, level: str, title: str, message: str):
    if callable(notify):
        notify(level, title, message)
        return
    if level == "error":
        messagebox.showerror(title, message)
    elif level == "warning":
        messagebox.showwarning(title, message)
    else:
        messagebox.showinfo(title, message)


def crear_pestana_buscar(tabs: ttk.Notebook, notify=None):
    frame = ttk.Frame(tabs)
    tabs.add(frame, text="Buscar / Editar")
    recibos_db.init_db()
    empty_state_var = tk.StringVar(value="Usá la búsqueda para listar recibos.")

    ttk.Label(frame, text="Buscar por:").grid(row=0, column=0, sticky="w", pady=5)
    criterio = ttk.Combobox(frame, values=["Número", "Cliente", "Fecha"], state="readonly")
    criterio.grid(row=0, column=1, padx=(0, 4))
    criterio.current(0)
    entrada = ttk.Entry(frame, width=34)
    entrada.grid(row=0, column=2)
    ttk.Button(frame, text="Buscar", command=lambda: _buscar()).grid(row=0, column=3, padx=6)

    filtros_fecha = ttk.Frame(frame)
    filtros_fecha.grid(row=1, column=0, columnspan=4, sticky="w", pady=(0, 4))
    ttk.Label(filtros_fecha, text="Desde (DD/MM/AAAA):").grid(row=0, column=0, padx=(0, 4))
    entrada_desde = ttk.Entry(filtros_fecha, width=12)
    entrada_desde.grid(row=0, column=1, padx=(0, 10))
    ttk.Label(filtros_fecha, text="Hasta (DD/MM/AAAA):").grid(row=0, column=2, padx=(0, 4))
    entrada_hasta = ttk.Entry(filtros_fecha, width=12)
    entrada_hasta.grid(row=0, column=3)

    resultados = tk.Listbox(frame, width=100, height=12)
    resultados.grid(row=2, column=0, columnspan=4, pady=10, sticky="we")
    ttk.Label(frame, textvariable=empty_state_var, style="Muted.TLabel").grid(row=3, column=0, columnspan=4, sticky="w")

    def _buscar():
        resultados.delete(0, tk.END)

        campo = criterio.get()
        valor_original = (entrada.get() or "").strip()
        valor = valor_original.lower()
        filas_por_num = {}

        fecha_desde_val = (entrada_desde.get() or "").strip()
        fecha_hasta_val = (entrada_hasta.get() or "").strip()
        if campo == "Fecha" and valor_original:
            fecha_desde_val = valor_original
            fecha_hasta_val = valor_original

        # 0) Consultar base de recibos
        try:
            filas_db = recibos_db.buscar_recibos(
                numero=valor_original if campo == "Número" else None,
                cliente=valor_original if campo == "Cliente" else None,
                fecha_desde=fecha_desde_val or None,
                fecha_hasta=fecha_hasta_val or None,
                limite=500,
            )
            for fila in filas_db:
                filas_por_num[fila["numero"]] = [
                    fila["numero"],
                    fila.get("cliente", ""),
                    fila.get("fecha", ""),
                    "",
                    f"{float(fila.get('total_bruto') or 0.0):.2f}",
                    "",
                ]
        except Exception:
            pass

        # 1) Buscar por PDFs en SALIDA_DIR (similar a pestaña Anular)
        try:
            pdfs = sorted(Path(SALIDA_DIR).glob("Recibo_*.pdf"))
            patron = re.compile(r"^Recibo_(\d{4}-\d{8})__(.+)\.pdf$", re.IGNORECASE)
            for p in pdfs:
                m = patron.match(p.name)
                if not m:
                    continue
                num, suf = m.groups()
                est = "Anulado" if suf.upper() == "ANULADO" else ""
                cli = "" if est else suf.replace("_", " ")

                campo_val = {"Número": num, "Cliente": cli, "Fecha": ""}.get(campo, p.name)
                if valor in str(campo_val).lower() or (campo == "Fecha" and not valor):
                    filas_por_num[num] = [num, cli, "", "", "", est]
        except Exception:
            pass

        # 2) Enriquecer/completar desde historial si disponible
        p = Path(HISTORIAL_PATH)
        if openpyxl is not None and p.exists():
            try:
                libro = openpyxl.load_workbook(p)
                hoja = libro.active
                for fila in hoja.iter_rows(min_row=2, values_only=True):
                    num = str(fila[0]) if fila and len(fila) > 0 else None
                    if not num:
                        continue
                    if num in filas_por_num:
                        filas_por_num[num] = list((fila or ())[:6])
                    else:
                        try:
                            campo_val = {"Número": fila[0], "Cliente": fila[1], "Fecha": fila[2]}.get(campo)
                            hay = valor in str(campo_val).lower()
                        except Exception:
                            hay = False
                        if hay or (not valor):
                            filas_por_num[num] = list((fila or ())[:6])
            except Exception:
                pass

        if not filas_por_num:
            empty_state_var.set("No hay resultados para esa búsqueda.")
            return
        empty_state_var.set(f"Se encontraron {len(filas_por_num)} resultado(s).")
        for fila6 in filas_por_num.values():
            resultados.insert(tk.END, " | ".join(str(c) for c in fila6))

    def editar():
        seleccion = resultados.get(tk.ACTIVE)
        if not seleccion:
            _notify(notify, "warning", "Seleccionar", "Seleccioná un recibo para editar.")
            return

        try:
            partes = [p.strip() for p in seleccion.split("|")]
            while len(partes) < 6:
                partes.append("")
            numero, cliente_row, fecha_row, subtotal_row, total_row, estado_row = partes[:6]
        except Exception:
            _notify(notify, "error", "Error", "No se pudo interpretar la fila seleccionada.")
            return

        # Intentar cargar JSON desde la base SQLite primero
        datos = {}
        try:
            fila_db = recibos_db.buscar_recibos(numero=numero, limite=1)
            if fila_db:
                datos = fila_db[0].get("datos") or {}
        except Exception:
            datos = {}

        # Si no se encontró en la base, intentar desde el historial
        if not datos and openpyxl is not None and Path(HISTORIAL_PATH).exists():
            try:
                libro = openpyxl.load_workbook(HISTORIAL_PATH)
                hoja = libro.active
                for row in hoja.iter_rows(min_row=2):
                    if str(row[0].value) == str(numero):
                        try:
                            c7 = hoja.cell(row=row[0].row, column=7).value
                            if c7:
                                datos = json.loads(c7)
                        except Exception:
                            datos = {}
                        break
            except Exception:
                datos = {}

        # Fallback: si no hay JSON en el Excel, intentar leer el PDF
        # original para precargar campos basicos.
        if not datos:
            try:
                from PyPDF2 import PdfReader
                base = Path(SALIDA_DIR)
                candidatos = sorted(base.glob(f"Recibo_{numero}__*.pdf"))
                if candidatos:
                    reader = PdfReader(str(candidatos[0]))
                    full_txt = "\n".join((p.extract_text() or "") for p in reader.pages)

                    def _find(pat, flags=0, default=""):
                        m = re.search(pat, full_txt, flags)
                        return m.group(1).strip() if m else default

                    def _to_float(s: str) -> float:
                        try:
                            s = str(s).strip()
                            if not s:
                                return 0.0
                            if s.isdigit():
                                return float(s)
                            if "," in s and "." not in s:
                                s = s.replace(".", "").replace(",", ".")
                            elif "," in s and "." in s:
                                if s.rfind(",") > s.rfind("."):
                                    s = s.replace(".", "").replace(",", ".")
                                else:
                                    s = s.replace(",", "")
                            else:
                                s = s.replace(",", "")
                            return float(s)
                        except Exception:
                            return 0.0

                    fecha     = _find(r"Fecha:\s*(\d{2}/\d{2}/\d{4})")
                    cliente   = _find(r"Cliente:\s*(.+)")
                    domicilio = _find(r"Domicilio:\s*(.+)")
                    localidad = _find(r"Localidad:\s*(.+)")
                    cuit      = _find(r"CUIT:\s*([\d\-\.\s]+)")
                    iva       = _find(r"Condici.?n\s+IVA:\s*(.+)")
                    total_s   = _find(r"Total:\s*\$?\s*([0-9\.,]+)")
                    total_v   = _to_float(total_s)

                    m = re.search(r"En concepto de:\s*(.*?)\s*Retenciones", full_txt, re.DOTALL)
                    concepto = m.group(1).strip() if m else ""

                    datos = {
                        "numero_recibo": numero,
                        "fecha": fecha or fecha_row,
                        "cliente": cliente or cliente_row,
                        "domicilio": domicilio,
                        "localidad": localidad,
                        "cuit": cuit,
                        "iva": iva,
                        "concepto": concepto,
                        "retenciones": {"Ganancias": 0.0, "SUSS": 0.0, "TEM": 0.0, "IIBB": 0.0},
                        "forma_pago": [],
                        "total": total_v or (total_row or "0"),
                    }
            except Exception:
                pass

        # Defaults si faltan
        datos.setdefault("numero_recibo", numero)
        datos.setdefault("fecha", fecha_row)
        datos.setdefault("cliente", cliente_row)
        datos.setdefault("domicilio", "")
        datos.setdefault("localidad", "")
        datos.setdefault("cuit", "")
        datos.setdefault("iva", "")
        datos.setdefault("concepto", "")
        datos.setdefault("retenciones", {"Ganancias": 0.0, "SUSS": 0.0, "TEM": 0.0, "IIBB": 0.0})
        datos.setdefault("forma_pago", [])
        datos.setdefault("total", total_row or "0")

        win = tk.Toplevel(frame)
        win.title(f"Editar recibo {numero}")

        campos = {}
        filas = [
            ("Número de recibo", "numero_recibo", True),
            ("Fecha (DD/MM/AAAA)", "fecha", False),
            ("Cliente", "cliente", False),
            ("Domicilio", "domicilio", False),
            ("Localidad", "localidad", False),
            ("CUIT", "cuit", False),
            ("Condición IVA", "iva", False),
            ("Total ($)", "total", False),
        ]
        for i, (label, key, ro) in enumerate(filas):
            ttk.Label(win, text=label).grid(row=i, column=0, sticky="e", padx=6, pady=3)
            entry = ttk.Entry(win, width=40)
            entry.grid(row=i, column=1, sticky="w", padx=6, pady=3)
            entry.insert(0, str(datos.get(key, "")))
            if ro:
                entry.config(state="readonly")
            campos[key] = entry

        # Estado
        idx_estado = len(filas)
        ttk.Label(win, text="Estado").grid(row=idx_estado, column=0, sticky="e", padx=6, pady=3)
        ent_estado = ttk.Combobox(win, values=["", "Anulado"], width=37, state="readonly")
        ent_estado.grid(row=idx_estado, column=1, sticky="w", padx=6, pady=3)
        ent_estado.set(estado_row or "")

        # Concepto
        ttk.Label(win, text="En concepto de").grid(row=idx_estado + 1, column=0, sticky="ne", padx=6)
        concepto_text = tk.Text(win, width=38, height=4)
        concepto_text.grid(row=idx_estado + 1, column=1, sticky="w", padx=6, pady=2)
        concepto_text.insert("1.0", datos.get("concepto", ""))

        # Retenciones
        base_row = idx_estado + 2
        ttk.Label(win, text="Retenciones ($)").grid(row=base_row, column=0, sticky="ne", padx=6)
        ret_frame = ttk.Frame(win)
        ret_frame.grid(row=base_row, column=1, sticky="w", padx=6, pady=4)
        ret_labels = ["Ganancias", "SUSS", "TEM", "IIBB"]
        ret_entries = {}
        for j, lbl in enumerate(ret_labels):
            ttk.Label(ret_frame, text=lbl).grid(row=0, column=j, padx=5)
        for j, lbl in enumerate(ret_labels):
            e = ttk.Entry(ret_frame, width=10)
            e.grid(row=1, column=j, padx=5)
            e.insert(0, str((datos.get("retenciones") or {}).get(lbl, 0)))
            ret_entries[lbl] = e

        # Total - Retenciones (solo lectura)
        ttk.Label(win, text="Total - Retenciones ($)").grid(row=base_row + 1, column=0, sticky="e", padx=6)
        total_neto_entry = ttk.Entry(win, width=40, state="readonly")
        total_neto_entry.grid(row=base_row + 1, column=1, sticky="w", padx=6, pady=2)

        def _parse_monetario_local(s) -> float:
            try:
                s = str(s).strip().replace(" ", "")
                if not s:
                    return 0.0
                if s.isdigit():
                    return float(s)
                if "," in s and "." in s:
                    if s.rfind(",") > s.rfind("."):
                        s = s.replace(".", "").replace(",", ".")
                    else:
                        s = s.replace(",", "")
                elif "," in s:
                    s = s.replace(".", "").replace(",", ".")
                else:
                    s = s.replace(",", "")
                return float(s)
            except Exception:
                return 0.0

        def recalcular_totales(*_):
            bruto = _parse_monetario_local(campos["total"].get())
            r_sum = sum(_parse_monetario_local(ret_entries[k].get()) for k in ret_labels)
            neto = max(0.0, bruto - r_sum)
            total_neto_entry.config(state="normal")
            total_neto_entry.delete(0, tk.END)
            total_neto_entry.insert(0, f"{neto:.2f}")
            total_neto_entry.config(state="readonly")

        campos["total"].bind("<KeyRelease>", recalcular_totales)
        for e in ret_entries.values():
            e.bind("<KeyRelease>", recalcular_totales)
        recalcular_totales()

        # Forma de pago
        ttk.Label(win, text="Forma de pago").grid(row=base_row + 2, column=0, sticky="ne", padx=6)
        fp_frame = ttk.Frame(win)
        fp_frame.grid(row=base_row + 2, column=1, sticky="w", padx=6, pady=4)
        for j, lbl in enumerate(["Tipo", "Número", "C/Banco", "Fecha", "Importe"]):
            ttk.Label(fp_frame, text=lbl).grid(row=0, column=j, padx=5)
        fp_tipo = ttk.Combobox(fp_frame, values=["Efectivo", "Cheque", "Transferencia"], width=12, state="readonly")
        fp_tipo.grid(row=1, column=0, padx=5)
        fp_nro = ttk.Entry(fp_frame, width=12);   fp_nro.grid(row=1, column=1, padx=5)
        fp_banco = ttk.Entry(fp_frame, width=16); fp_banco.grid(row=1, column=2, padx=5)
        fp_fecha = ttk.Entry(fp_frame, width=12); fp_fecha.grid(row=1, column=3, padx=5)
        fp_importe = ttk.Entry(fp_frame, width=12); fp_importe.grid(row=1, column=4, padx=5)

        pagos_tree = ttk.Treeview(fp_frame, columns=("tipo", "numero", "banco", "fecha", "importe"), show="headings", height=6)
        for col, txt, w in [("tipo", "Tipo", 120), ("numero", "Número", 150), ("banco", "C/Banco", 150), ("fecha", "Fecha", 90), ("importe", "Importe", 110)]:
            pagos_tree.heading(col, text=txt)
            pagos_tree.column(col, width=w, anchor="w")
        pagos_tree.grid(row=2, column=0, columnspan=5, sticky="ew", padx=2, pady=(6, 2))

        for fp in (datos.get("forma_pago") or []):
            try:
                pagos_tree.insert("", "end", values=(
                    fp.get("tipo", ""), fp.get("numero", ""), fp.get("banco", ""), fp.get("fecha", ""), f"{_parse_monetario_local(fp.get('importe', 0)):.2f}"
                ))
            except Exception:
                continue

        def agregar_pago():
            t = fp_tipo.get().strip()
            n = fp_nro.get().strip()
            b = fp_banco.get().strip()
            f = fp_fecha.get().strip()
            im = _parse_monetario_local(fp_importe.get().strip())
            if not t:
                _notify(notify, "error", "Pago", "Seleccioná un tipo.")
                return
            if im <= 0:
                _notify(notify, "error", "Pago", "Importe inválido.")
                return
            pagos_tree.insert("", "end", values=(t, n, b, f, f"{im:.2f}"))

        def quitar_pago():
            sel = pagos_tree.selection()
            for iid in sel:
                pagos_tree.delete(iid)

        btns_fp = ttk.Frame(fp_frame)
        btns_fp.grid(row=3, column=0, columnspan=5, sticky="w", pady=(2, 0))
        ttk.Button(btns_fp, text="Agregar pago", command=agregar_pago).grid(row=0, column=0, padx=(0, 6))
        ttk.Button(btns_fp, text="Quitar seleccionado", command=quitar_pago).grid(row=0, column=1)

        def _sanitizar_nombre_cliente(s: str) -> str:
            return (s or "Cliente").replace(" ", "_")

        def _ruta_destino(num: str, cli: str, est: str) -> Path:
            base = Path(SALIDA_DIR)
            if (est or "").strip().lower() == "anulado":
                return base / f"Recibo_{num}__ANULADO.pdf"
            return base / f"Recibo_{num}__{_sanitizar_nombre_cliente(cli)}.pdf"

        def _colectar_pagos():
            pagos = []
            for iid in pagos_tree.get_children():
                t, n, b, f, im = pagos_tree.item(iid, "values")
                pagos.append({"tipo": t, "numero": n, "banco": b, "fecha": f, "importe": _parse_monetario_local(im)})
            return pagos

        def guardar():
            try:
                num = campos["numero_recibo"].get().strip()
                fec = campos["fecha"].get().strip()
                cli = campos["cliente"].get().strip()
                dom = campos["domicilio"].get().strip()
                loc = campos["localidad"].get().strip()
                cuit = campos["cuit"].get().strip()
                iva = campos["iva"].get().strip()
                bruto = _parse_monetario_local(campos["total"].get())
                est = ent_estado.get().strip()

                if not fec:
                    _notify(notify, "error", "Error", "La fecha es obligatoria.")
                    return
                if not cli:
                    _notify(notify, "error", "Error", "El cliente es obligatorio.")
                    return

                ret = {k: _parse_monetario_local(ret_entries[k].get()) for k in ret_labels}
                ret_sum = sum(ret.values())
                fps = _colectar_pagos()

                suma_fp = sum(p.get("importe", 0.0) for p in fps)
                if abs(suma_fp - (bruto - ret_sum)) > 0.01:
                    if not messagebox.askyesno(
                        "Atención",
                        f"La suma de pagos (${suma_fp:.2f}) no coincide con (Total - Retenciones) (${(bruto - ret_sum):.2f}).\n¿Continuar?"
                    ):
                        return

                datos_actual = {
                    "numero_recibo": num,
                    "fecha": fec,
                    "cliente": cli,
                    "domicilio": dom,
                    "localidad": loc,
                    "cuit": cuit,
                    "iva": iva,
                    "concepto": concepto_text.get("1.0", tk.END).strip(),
                    "retenciones": ret,
                    "forma_pago": fps,
                    "total": bruto,
                }

                qr_payload = build_qr_data(datos_actual, BASE_QR_URL, QR_SECRET_KEY)
                destino = _ruta_destino(num, cli, est)
                destino.parent.mkdir(parents=True, exist_ok=True)

                generar_pdf(
                    datos=datos_actual,
                    ruta_salida=destino,
                    logo_path=None,
                    firma_path=str(ASSETS_DIR / "firma.png"),
                    anulado=(est or "").lower() == "anulado",
                    qr_data=qr_payload,
                    template_pdf=str(ASSETS_DIR / "MODELO 2.pdf"),
                )

                # Limpiar PDFs anteriores con mismo número
                try:
                    base = Path(SALIDA_DIR)
                    for p in base.glob(f"Recibo_{num}__*.pdf"):
                        if p.resolve() != destino.resolve():
                            p.unlink(missing_ok=True)
                except Exception:
                    pass

                # Guardar historial con JSON (si es posible)
                try:
                    upsert_historial_con_json(
                        numero=num,
                        cliente=cli,
                        fecha=fec,
                        subtotal="",
                        total=bruto,
                        estado=est,
                        datos=datos_actual,
                    )
                except Exception:
                    # Si falta openpyxl o hubo error, solo avisar, el PDF ya está regenerado
                    _notify(notify, "warning", "Historial", "No se pudo actualizar el historial (openpyxl no instalado o archivo bloqueado).")

                # Refrescar lista
                nueva_fila = " | ".join([num, cli, fec, "", f"{bruto}", est])
                sel_idx = resultados.curselection()
                if sel_idx:
                    resultados.delete(sel_idx[0])
                    resultados.insert(sel_idx[0], nueva_fila)

                _notify(notify, "success", "Listo", "Recibo actualizado y PDF regenerado.")
                win.destroy()
            except Exception as e:
                _notify(notify, "error", "Error", str(e))

        btns = ttk.Frame(win)
        btns.grid(row=base_row + 3, column=0, columnspan=2, pady=8)
        ttk.Button(btns, text="Guardar", command=guardar).grid(row=0, column=0, padx=6)
        ttk.Button(btns, text="Cancelar", command=win.destroy).grid(row=0, column=1, padx=6)

    def busqueda_global(texto: str):
        txt = (texto or "").strip()
        if not txt:
            return
        if re.match(r"^\d{2}/\d{2}/\d{4}$", txt):
            criterio.set("Fecha")
        elif re.match(r"^\d{4}-\d{8}$", txt):
            criterio.set("Número")
        else:
            criterio.set("Cliente")
        entrada.delete(0, tk.END)
        entrada.insert(0, txt)
        _buscar()

    ttk.Button(frame, text="Editar seleccionado", command=editar).grid(row=4, column=0, columnspan=1, pady=8, sticky="w")
    return {"frame": frame, "buscar_global": busqueda_global, "buscar_local": _buscar}
